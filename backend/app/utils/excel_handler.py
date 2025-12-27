"""
Excel导入导出工具类
支持数据的批量导入和导出
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Excel处理工具类"""

    # 样式定义
    HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    CELL_FONT = Font(name='微软雅黑', size=10)
    CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

    BORDER_THIN = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        columns: Dict[str, str],
        sheet_name: str = 'Sheet1',
        title: Optional[str] = None
    ) -> BytesIO:
        """
        导出数据到Excel

        Args:
            data: 数据列表，每个元素是一个字典
            columns: 列定义，key是数据字段名，value是Excel列标题
            sheet_name: 工作表名称
            title: 可选的标题（会添加在第一行）

        Returns:
            BytesIO: Excel文件的字节流

        Example:
            data = [
                {'id': 1, 'name': '客户A', 'phone': '13800138000'},
                {'id': 2, 'name': '客户B', 'phone': '13900139000'}
            ]
            columns = {'id': 'ID', 'name': '客户名称', 'phone': '联系电话'}
            excel_file = ExcelHandler.export_to_excel(data, columns)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 起始行（如果有标题，从第2行开始）
        start_row = 1
        if title:
            # 添加标题行
            ws.merge_cells(f'A1:{get_column_letter(len(columns))}1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(name='微软雅黑', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            start_row = 2

        # 写入表头
        for col_idx, (field, header) in enumerate(columns.items(), 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = ExcelHandler.HEADER_FONT
            cell.fill = ExcelHandler.HEADER_FILL
            cell.alignment = ExcelHandler.HEADER_ALIGNMENT
            cell.border = ExcelHandler.BORDER_THIN

            # 设置列宽（根据标题长度）
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(15, len(header) * 2 + 2)

        # 设置表头行高
        ws.row_dimensions[start_row].height = 25

        # 写入数据
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, field in enumerate(columns.keys(), 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(field)

                # 格式化数据
                if value is None:
                    cell.value = ''
                elif isinstance(value, Decimal):
                    cell.value = float(value)
                elif isinstance(value, datetime):
                    cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    cell.value = '是' if value else '否'
                else:
                    cell.value = str(value)

                cell.font = ExcelHandler.CELL_FONT
                cell.alignment = ExcelHandler.CELL_ALIGNMENT
                cell.border = ExcelHandler.BORDER_THIN

        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_from_excel(
        file: BytesIO,
        columns: Dict[str, str],
        sheet_name: Optional[str] = None,
        start_row: int = 2
    ) -> List[Dict[str, Any]]:
        """
        从Excel导入数据

        Args:
            file: Excel文件的字节流
            columns: 列定义，key是数据字段名，value是Excel列标题
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            start_row: 数据起始行（默认第2行，第1行是表头）

        Returns:
            List[Dict]: 导入的数据列表

        Raises:
            ValueError: 如果Excel格式不正确

        Example:
            columns = {'id': 'ID', 'name': '客户名称', 'phone': '联系电话'}
            data = ExcelHandler.import_from_excel(file, columns)
        """
        try:
            wb = load_workbook(file)
            ws = wb[sheet_name] if sheet_name else wb.active

            # 读取表头（第1行或start_row-1行）
            header_row = start_row - 1
            headers = {}
            for col_idx, cell in enumerate(ws[header_row], 1):
                header_value = str(cell.value).strip() if cell.value else ''
                # 找到对应的字段名
                for field, header in columns.items():
                    if header == header_value:
                        headers[col_idx] = field
                        break

            if not headers:
                raise ValueError("Excel表头与预期格式不匹配")

            # 读取数据
            result = []
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                # 跳过空行
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                row_data = {}
                for col_idx, field in headers.items():
                    value = row[col_idx - 1] if col_idx <= len(row) else None

                    # 数据清洗
                    if value is not None:
                        if isinstance(value, str):
                            value = value.strip()
                            if value == '':
                                value = None

                    row_data[field] = value

                if row_data:  # 如果行数据不为空
                    result.append(row_data)

            return result

        except Exception as e:
            raise ValueError(f"Excel文件解析失败: {str(e)}")

    @staticmethod
    def create_template(
        columns: Dict[str, str],
        sheet_name: str = 'Sheet1',
        title: Optional[str] = None,
        sample_data: Optional[List[Dict[str, Any]]] = None
    ) -> BytesIO:
        """
        创建Excel导入模板

        Args:
            columns: 列定义
            sheet_name: 工作表名称
            title: 可选标题
            sample_data: 可选的示例数据

        Returns:
            BytesIO: Excel模板文件
        """
        # 如果没有示例数据，创建一行空数据作为示例
        if sample_data is None:
            sample_data = [{}]

        return ExcelHandler.export_to_excel(sample_data, columns, sheet_name, title)

    @staticmethod
    def validate_data(
        data: List[Dict[str, Any]],
        required_fields: List[str],
        validators: Optional[Dict[str, callable]] = None
    ) -> tuple[bool, List[str]]:
        """
        验证导入的数据

        Args:
            data: 待验证的数据
            required_fields: 必填字段列表
            validators: 自定义验证器字典，key是字段名，value是验证函数

        Returns:
            tuple: (是否验证通过, 错误信息列表)

        Example:
            validators = {
                'phone': lambda x: len(x) == 11 if x else True,
                'email': lambda x: '@' in x if x else True
            }
            is_valid, errors = ExcelHandler.validate_data(data, ['name'], validators)
        """
        errors = []

        for idx, row in enumerate(data, 1):
            # 检查必填字段
            for field in required_fields:
                if not row.get(field):
                    errors.append(f"第{idx}行：{field}字段不能为空")

            # 自定义验证
            if validators:
                for field, validator in validators.items():
                    value = row.get(field)
                    try:
                        if not validator(value):
                            errors.append(f"第{idx}行：{field}字段格式不正确")
                    except Exception as e:
                        errors.append(f"第{idx}行：{field}字段验证失败 - {str(e)}")

        return len(errors) == 0, errors


class DataFrameExporter:
    """
    使用Pandas进行更复杂的数据导出
    适用于需要数据处理、统计、多Sheet等场景
    """

    @staticmethod
    def export_multi_sheet(
        data_dict: Dict[str, pd.DataFrame],
        file_name: Optional[str] = None
    ) -> BytesIO:
        """
        导出多个Sheet到一个Excel文件

        Args:
            data_dict: 字典，key是sheet名称，value是DataFrame
            file_name: 可选的文件名

        Returns:
            BytesIO: Excel文件
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        output.seek(0)
        return output

    @staticmethod
    def export_with_charts(
        df: pd.DataFrame,
        sheet_name: str = 'Data',
        chart_config: Optional[Dict] = None
    ) -> BytesIO:
        """
        导出带图表的Excel

        Args:
            df: 数据DataFrame
            sheet_name: 工作表名称
            chart_config: 图表配置（未来扩展）

        Returns:
            BytesIO: Excel文件
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # TODO: 添加图表功能

        output.seek(0)
        return output
